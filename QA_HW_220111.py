from os import system
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


options = Options()
options.add_argument("--disable-notifications")
chrome = webdriver.Chrome('./chromedriver', chrome_options=options)
chrome.maximize_window()
url = 'https://www.cathaybk.com.tw/cathaybk/personal/investment/etf/search/?hotissue=05'
response = chrome.get(url)
for x in range(0, 2):
    chrome.execute_script("window.scrollTo(0,document.body.scrollHeight)")
    time.sleep(5)


table = chrome.find_element(By.ID,"hasResult")
thread = table.find_element(By.CLASS_NAME,"cubinvest-l-table__thead")
titles = thread.find_elements(By.XPATH, './/div[@class="cubinvest-l-table__th"][not(@style="display: none;")]')
titles.pop(0)
titles.pop()
body = table.find_element(By.ID,"resultContainer")
products = body.find_elements(By.XPATH, './/div[@class="cubinvest-l-table__tr"]')


workbook = openpyxl.Workbook()
sheet = workbook.worksheets[0]
sheet.row_dimensions[1].height = 20
sheet.column_dimensions['B'].width = 36
sheet.column_dimensions['C'].width = 12

for (idx, title) in enumerate(titles):
    sheet.cell(row=1, column=idx+1).alignment = Alignment(horizontal="center")
    sheet.cell(row=1, column=idx+1).font = Font(bold=True)
    sheet.cell(row=1, column=idx+1).value = title.text

for (idx_row, product) in enumerate(products):
    product_infos = product.find_elements(By.XPATH, './/div[contains(@class, "cubinvest-l-table__td")][not(@style="display: none;")][not(@style="display:none;")]')
    product_infos.pop(0)
    product_infos.pop()
    for(idx_col, info) in enumerate(product_infos):
        sheet.cell(row=idx_row+2, column=idx_col+1).alignment = Alignment(horizontal="center",vertical="center") if (idx_col!=1) else Alignment(vertical="center")
        if info.text.startswith("-"):
            sheet.cell(row=idx_row+2, column=idx_col+1).font = Font(color='ff0000') 
        sheet.cell(row=idx_row+2, column=idx_col+1).value = info.text

workbook.save('CUB_ETF.xlsx')
workbook.close()